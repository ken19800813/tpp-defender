import json
import os
import base64
import uuid
import requests
from dataclasses import dataclass, asdict
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Excel 規則範本欄位順序（下載範本與上傳解析都依照此順序，改動時兩邊要同步）
EXCEL_HEADERS = ["關鍵字", "回應文1", "回應文2", "回應文3", "是否為最優先規則", "是否分享到社群"]

# GitHub URL Base64編碼（防止易被解讀）- ken19800813/tpp-defender
REMOTE_SECURITY_URL_B64 = "aHR0cHM6Ly9yYXcuZ2l0aHVidXNlcmNvbnRlbnQuY29tL2tlbjE5ODAwODEzL3RwcC1kZWZlbmRlci9tYWluL3NlY3VyaXR5X3J1bGVzLmpzb24="
REMOTE_SECURITY_URL = base64.b64decode(REMOTE_SECURITY_URL_B64).decode()
CACHE_FILE = "security_cache.json"
VERSION_INFO_FILE = "security_version.json"

# 黑名單頻道、防禦關鍵字規則、跑馬燈文字改由 LINEBOT 後台的
# 「直播小幫手」獨立模組統一編輯與更新(/ken_admin/livestream/*)，
# GitHub security_rules.json 只再保留 forbidden_attack_words 一項。
LIVESTREAM_CONFIG_API_URL = "https://line-news-0p7m.onrender.com/api/livestream/config"
LIVESTREAM_VERSION_API_URL = "https://line-news-0p7m.onrender.com/api/livestream/version"
LIVESTREAM_CACHE_FILE = "livestream_cache.json"

# 廣告推播：獨立於上面的設定同步之外，用比較短的輪詢間隔單獨檢查，
# 避免廣告要等好幾分鐘才生效，也不用每次都把整包規則資料抓一次。
LIVESTREAM_ADS_API_URL = "https://line-news-0p7m.onrender.com/api/livestream/ads/active"
SEEN_ADS_FILE = "seen_ads.json"

# 使用者自願分享自訂規則到社群資料庫的投遞端點；後端會獨立再驗一次內容
# （黑名單/速率限制），本機驗證只是第一道防線。
LIVESTREAM_SHARE_RULE_API_URL = "https://line-news-0p7m.onrender.com/api/livestream/rules/share"


@dataclass
class Rule:
    id: str
    trigger_keywords: List[str]
    match_type: str
    reply_pool: List[str]
    is_enabled: bool = True
    is_priority: bool = False  # 最優先回覆規則：命中任何規則的關鍵字後，
    # 若存在啟用中的優先規則，回覆內容一律改用該規則，不管實際命中的是哪一條
    wants_share: bool = False      # 使用者是否勾選要分享這條規則到社群資料庫
    already_shared: bool = False   # 目前這條規則的內容是否已成功上傳過（編輯後會重置）


@dataclass
class SafetySettings:
    user_data_dir: str = "./user_data"


class ConfigManager:
    def __init__(self, filepath="config.json"):
        self.filepath = filepath
        self.settings = SafetySettings()
        self.rules: List[Rule] = []
        self.user_rules: List[Rule] = []
        self.forbidden_words: List[str] = []
        self.default_rules_data: List[dict] = []
        self.locked_channels: List[str] = []
        self.marquee_messages: List[str] = []
        self.marquee_speed_level: int = 4
        self.auto_send_enabled: bool = False
        # 頻道主/版主模式：開啟後側翼攻擊彈窗會出現「封鎖此人」按鈕。
        # 實際能不能封鎖成功仍取決於 YouTube 帳號本身是否有本頻道的板主
        # 權限——這個開關只是「要不要顯示按鈕」，真正的權限判斷發生在
        # 按下按鈕當下（bot_engine._try_ban 嘗試操作，選單沒有封鎖選項
        # 就代表沒權限，直接回報失敗，不會誤導使用者）。
        self.moderator_mode: bool = False
        self.last_share_date = None  # 每日分享批次上次成功嘗試的日期字串 "YYYY-MM-DD"，None 代表從未執行過
        self.seen_ad_ids = self._load_seen_ads()
        self.synced_config_version = None  # 本機目前已套用的直播設定 version，None=從未成功下載過

        self.fetch_remote_rules()
        self.load_livestream_config_or_cache()
        self.load()
        self._rebuild_rules()

    def load_livestream_config_or_cache(self):
        """啟動時的省流策略：優先用本機快取秒開，只有首次安裝（無快取）
        才做一次 bootstrap 全量下載。之後的更新改由前端主動偵測版本、
        使用者確認後再呼叫 fetch_livestream_config()。"""
        if os.path.exists(LIVESTREAM_CACHE_FILE):
            self._load_livestream_from_local_cache()
        else:
            self.fetch_livestream_config()

    def fetch_livestream_version(self):
        """極輕量版本查詢，成功回傳 int，失敗一律回 None（吞例外）"""
        try:
            res = requests.get(LIVESTREAM_VERSION_API_URL, timeout=4)
            if res.status_code != 200:
                return None
            data = res.json()
            if not data.get("success"):
                return None
            version = data.get("version")
            if isinstance(version, int):
                return version
            return None
        except Exception:
            return None

    def fetch_remote_rules(self):
        """智慧 ETag 版控：有更動才下載，否則使用本機快取秒開
        關閉 gzip 協商是為了避開 raw.githubusercontent.com 的 CDN 對
        gzip 壓縮版本的快取變體，該變體在剛 push 後常有數分鐘的延遲，
        會導致抓到比 curl 看到的還舊的內容。檔案本身很小，不壓縮也無影響。"""
        headers = {"Accept-Encoding": "identity"}
        if os.path.exists(VERSION_INFO_FILE):
            try:
                with open(VERSION_INFO_FILE, "r", encoding="utf-8") as f:
                    version_data = json.load(f)
                    if "ETag" in version_data:
                        headers["If-None-Match"] = version_data["ETag"]
            except Exception:
                pass

        try:
            res = requests.get(REMOTE_SECURITY_URL, headers=headers, timeout=4)
            if res.status_code == 304:
                self._load_from_local_cache()
                return
            elif res.status_code == 200:
                data = res.json()
                self.forbidden_words = data.get("forbidden_attack_words", [])

                with open(CACHE_FILE, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)

                with open(VERSION_INFO_FILE, "w", encoding="utf-8") as f:
                    version_info = {}
                    if "ETag" in res.headers:
                        version_info["ETag"] = res.headers["ETag"]
                    if "Last-Modified" in res.headers:
                        version_info["Last-Modified"] = res.headers["Last-Modified"]
                    json.dump(version_info, f, indent=4)
                return
        except Exception:
            pass

        self._load_from_local_cache()

    def fetch_livestream_config(self):
        """黑名單頻道、防禦關鍵字規則、跑馬燈文字統一從 LINEBOT 後台的
        「直播小幫手」模組同步。連線失敗時改用本機快取，避免整個空白。"""
        try:
            res = requests.get(LIVESTREAM_CONFIG_API_URL, timeout=4)
            if res.status_code == 200:
                data = res.json()
                if data.get("success"):
                    self.locked_channels = data.get("locked_channels", [])
                    self.default_rules_data = data.get("default_defense_rules", [])
                    self.marquee_messages = data.get("marquee_messages", [])
                    self.marquee_speed_level = data.get("marquee_speed_level", 4)
                    version = data.get("version")
                    if isinstance(version, int):
                        self.synced_config_version = version
                    # 額外把版本號寫進 cache（用 _synced_version 這個 key，
                    # 避免跟後端未來可能新增的 version 欄位共用同一份 payload 時混淆）
                    cache_payload = dict(data)
                    cache_payload["_synced_version"] = self.synced_config_version
                    with open(LIVESTREAM_CACHE_FILE, "w", encoding="utf-8") as f:
                        json.dump(cache_payload, f, ensure_ascii=False, indent=4)
                    return
        except Exception:
            pass

        self._load_livestream_from_local_cache()

    def _load_livestream_from_local_cache(self):
        if os.path.exists(LIVESTREAM_CACHE_FILE):
            try:
                with open(LIVESTREAM_CACHE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.locked_channels = data.get("locked_channels", [])
                    self.default_rules_data = data.get("default_defense_rules", [])
                    self.marquee_messages = data.get("marquee_messages", [])
                    self.marquee_speed_level = data.get("marquee_speed_level", 4)
                    self.synced_config_version = data.get("_synced_version")
            except Exception:
                pass

    def _load_seen_ads(self):
        if os.path.exists(SEEN_ADS_FILE):
            try:
                with open(SEEN_ADS_FILE, "r", encoding="utf-8") as f:
                    return set(json.load(f))
            except Exception:
                return set()
        return set()

    def _save_seen_ads(self):
        try:
            with open(SEEN_ADS_FILE, "w", encoding="utf-8") as f:
                json.dump(list(self.seen_ad_ids), f)
        except Exception:
            pass

    def fetch_new_ads(self):
        """回傳目前有效、但這台電腦還沒顯示過的廣告清單，並立即標記為已顯示
        （避免同一支廣告在下次輪詢時又跳出來一次）。連線失敗時安靜回傳空清單，
        不影響其他功能。"""
        try:
            res = requests.get(LIVESTREAM_ADS_API_URL, timeout=4)
            if res.status_code != 200:
                return []
            data = res.json()
            if not data.get("success"):
                return []
            ads = data.get("ads", [])
        except Exception:
            return []

        new_ads = [ad for ad in ads if ad.get("id") not in self.seen_ad_ids]
        if new_ads:
            for ad in new_ads:
                self.seen_ad_ids.add(ad["id"])
            self._save_seen_ads()
        return new_ads

    def _load_from_local_cache(self):
        """從本機快取載入 GitHub 端規則（forbidden_attack_words）"""
        if os.path.exists(CACHE_FILE):
            try:
                with open(CACHE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.forbidden_words = data.get("forbidden_attack_words", [])
            except Exception:
                pass

    def load(self):
        """載入本機設定（自訂規則、全自動送出開關等個人偏好，不含雲端預設規則）"""
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    d = json.load(f)
                    self.user_rules = [Rule(**r) for r in d.get("rules", [])]
                    self.auto_send_enabled = d.get("auto_send_enabled", False)
                    self.moderator_mode = d.get("moderator_mode", False)
                    self.last_share_date = d.get("last_share_date")
            except Exception:
                self.user_rules = []
        else:
            self.user_rules = []

    def _rebuild_rules(self):
        """合併雲端預設規則與使用者自訂規則，雲端規則更新時會自動反映在這裡"""
        cloud_rules = [Rule(**r) for r in self.default_rules_data]
        self.rules = cloud_rules + self.user_rules

    def validate_custom_rule(self, user_keywords: List[str], user_replies: List[str]) -> tuple:
        """雙向審查：若輸入內容包含中央禁用詞，則攔截拒絕儲存。
        回傳 (is_valid: bool, error_message: str)；valid 時 error_message 為空"""
        for word in self.forbidden_words:
            for kw in user_keywords:
                if word in kw:
                    return False, f"關鍵字不能包含：{word}"
            for rp in user_replies:
                if word in rp:
                    return False, f"回應文不能包含：{word}"
        return True, ""

    def share_rule_to_cloud(self, trigger_keywords: List[str],
                             reply_pool: List[str]) -> tuple:
        """把使用者自訂規則投遞到雲端社群資料庫，讓其他使用者也能同步下載到。
        後端會獨立再驗一次（黑名單/速率限制），驗證失敗時 error 欄位會回一段
        中文說明；連線失敗/timeout/JSON parse 失敗一律吞掉並回統一提示，
        呼叫端不需要 catch。回傳 (success, error_message)。"""
        payload = {
            "trigger_keywords": trigger_keywords,
            "reply_pool": reply_pool,
        }
        try:
            res = requests.post(
                LIVESTREAM_SHARE_RULE_API_URL, json=payload, timeout=4
            )
            try:
                data = res.json()
            except Exception:
                return (False, "伺服器回應格式異常，稍後可再試")

            if res.status_code == 200 and data.get("success"):
                return (True, "")
            error_msg = data.get("error") or "伺服器暫時無法處理，稍後可再試"
            return (False, error_msg)
        except Exception:
            return (False, "網路連線失敗，稍後可再試")

    def save(self):
        """存檔本機設定（自訂規則、全自動送出開關，雲端預設規則不寫入本機檔案）"""
        os.makedirs(os.path.dirname(self.filepath) or ".", exist_ok=True)
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "rules": [asdict(r) for r in self.user_rules],
                    "auto_send_enabled": self.auto_send_enabled,
                    "moderator_mode": self.moderator_mode,
                    "last_share_date": self.last_share_date,
                },
                f,
                indent=4,
                ensure_ascii=False
            )

    def set_auto_send_enabled(self, enabled: bool):
        """開關全自動送出模式，並持久化到本機設定檔"""
        self.auto_send_enabled = enabled

    def set_moderator_mode(self, enabled: bool):
        """開關頻道主/版主模式（決定側翼攻擊彈窗是否顯示封鎖按鈕），持久化"""
        self.moderator_mode = enabled
        self.save()

    def export_rule_template(self, save_path: str):
        """產生防禦規則 Excel 範本，帶表頭與範例列，方便使用者照格式填寫。
        欄位順序固定為 EXCEL_HEADERS，跟 import_rules_from_excel() 的解析順序
        必須保持一致，改任一邊都要同步改另一邊。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "防禦規則"
        ws.append(EXCEL_HEADERS)
        ws.append(["貪污犯,貪污,索賄", "講到貪污，那你一定很討厭貪污的人吧", "", "", "否", "否"])
        ws.append(["檳榔,哭文哲", "這句話跟事實不符喔", "我們可以理性討論", "", "否", "否"])
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 30

        # 「是否為最優先規則」「是否分享到社群」欄位（E、F）改成下拉選單，
        # 只能選「是」/「否」，避免使用者手動輸入時打錯字（例如「Y」「有」）
        # 導致 import_rules_from_excel() 用 == "是" 判斷時誤判成否。
        # 範圍留到第 1000 列，讓使用者自己新增列時也還在下拉選單涵蓋範圍內。
        yes_no_validation_e = DataValidation(
            type="list", formula1='"是,否"', allow_blank=True, showDropDown=False
        )
        yes_no_validation_f = DataValidation(
            type="list", formula1='"是,否"', allow_blank=True, showDropDown=False
        )
        ws.add_data_validation(yes_no_validation_e)
        ws.add_data_validation(yes_no_validation_f)
        yes_no_validation_e.add(f"E2:E1000")
        yes_no_validation_f.add(f"F2:F1000")

        wb.save(save_path)

    def import_rules_from_excel(self, file_path: str) -> List[dict]:
        """讀取使用者填好的 Excel，逐列解析成規則候選清單，回傳給呼叫端
        （main.py）逐筆做髒話檢查、UI 確認、決定是否分享後再真正呼叫
        add_rule()。這裡只負責解析格式，不寫入、不驗證禁用詞——驗證是
        呼叫端的責任，因為呼叫端才知道要怎麼呈現錯誤給使用者看。

        回傳格式：每列一個 dict：
        {
            "row": 該列在 Excel 的實際行號(從2開始，1是表頭)，
            "keywords": [str,...],
            "replies": [str,...],
            "is_priority": bool,
            "wants_share": bool,
            "error": str 或 None（若解析失敗，keywords/replies 會是空的）
        }
        空白列（關鍵字與所有回應文皆空）直接略過，不放進回傳清單。"""
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        results = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            padded = list(row) + [None] * (len(EXCEL_HEADERS) - len(row))
            raw_keywords, raw_r1, raw_r2, raw_r3, raw_priority, raw_share = padded[:6]

            keywords_str = str(raw_keywords).strip() if raw_keywords else ""
            replies = [str(r).strip() for r in (raw_r1, raw_r2, raw_r3) if r and str(r).strip()]

            if not keywords_str and not replies:
                continue  # 整列空白，略過不當錯誤

            keywords = [k.strip() for k in keywords_str.split(",") if k.strip()]
            is_priority = str(raw_priority).strip() == "是" if raw_priority else False
            wants_share = str(raw_share).strip() == "是" if raw_share else False

            error = None
            if not keywords and not is_priority:
                error = "關鍵字欄位是空的"
            elif not replies:
                error = "至少需要填一句回應文"

            results.append({
                "row": row_idx,
                "keywords": keywords,
                "replies": replies,
                "is_priority": is_priority,
                "wants_share": wants_share,
                "error": error,
            })
        wb.close()
        return results

    def add_rule(self, rule: Rule):
        """新增使用者自訂規則"""
        self.user_rules.append(rule)
        self._rebuild_rules()
        self.save()

    def delete_rule(self, rule_id: str):
        """刪除使用者自訂規則（雲端預設規則無法在此刪除）"""
        self.user_rules = [r for r in self.user_rules if r.id != rule_id]
        self._rebuild_rules()
        self.save()

    def update_rule(self, rule_id: str, updated_rule: Rule):
        """更新使用者自訂規則"""
        self.user_rules = [updated_rule if r.id == rule_id else r for r in self.user_rules]
        self._rebuild_rules()
        self.save()

    def reorder_user_rules(self, ordered_ids: List[str]):
        """依 ordered_ids 的順序重排 self.user_rules。
        找不到對應 id 的忽略；self.user_rules 裡有但不在 ordered_ids 內的規則，
        維持原相對順序附加到最後（防禦性處理，避免資料遺失）。"""
        by_id = {r.id: r for r in self.user_rules}
        seen = set()
        new_rules: List[Rule] = []
        for rid in ordered_ids:
            if rid in by_id and rid not in seen:
                new_rules.append(by_id[rid])
                seen.add(rid)
        for r in self.user_rules:
            if r.id not in seen:
                new_rules.append(r)
                seen.add(r.id)
        self.user_rules = new_rules
        self._rebuild_rules()
        self.save()
